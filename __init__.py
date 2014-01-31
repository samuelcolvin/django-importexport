import openpyxl
from django.db import models as db_models
from django.core.exceptions import ObjectDoesNotExist
import settings

default_imex_fields = ['xl_id', 'name', 'description', 'comment']

class ImportExtra:
    def __init__(self, ws, headings):
        self._ws = ws
        self._headings = headings
    
    def get_row(self, *args, **kwargs):
        pass
        
class ImportM2MBase(ImportExtra):
    def __init__(self, *args, **kwargs):
        ImportExtra.__init__(self, *args, **kwargs)
            
    def get_row(self, main_item, row):
        main_i_field = getattr(main_item, self.m2m_field_name)
        m2m_model = main_i_field.model
        for comp_col in self._headings[self.m2m_field_name]:
            value = self._ws.cell(row=row, column=comp_col).value
            if value is not None:
                try:
                    main_i_field.add(m2m_model.objects.get(xl_id=value))
                except ObjectDoesNotExist:
                    raise Exception('ERROR: item with xl_id = %d does not exist in %s' % (value, m2m_model.__name__))
        
class ExportExtra(object):
    def __init__(self, ws, firstcol):
        self._ws = ws
        self._firstcol = firstcol
        
    def add_headings(self, row):
        pass
    
    def add_row(self, main_item, row):
        pass

class RedExtra(ExportExtra):
    print_color = 'RED'
    def __init__(self, ws, firstcol):
        self._red_columns = [i + firstcol for i in range(len(self.lookups))]
        ExportExtra.__init__(self, ws, firstcol)
        
    def add_headings(self, row):
        for index, lookup in enumerate(self.lookups):
            c = self._ws.cell(row = row, column=self._red_columns[index])
            c.value = lookup['heading']
            c.style.font.bold = True
            self.set_red(c)
            if 'func' not in lookup and 'ref_col' in lookup:
                get_col='B'
                if 'get_col' in lookup:
                    get_col = lookup['get_col']
                self.add_1000_formula(row + 1, self._red_columns[index], lookup['ref_col'], lookup['sheet'], get_col)
        
    def add_row(self, main_item, row):
        for index, lookup in enumerate(self.lookups):
            if 'func' in lookup:
                self.add_red_value(main_item, row, self._red_columns[index], func = lookup['func'])
            elif 'field' in lookup:
                self.add_red_value(main_item, row, self._red_columns[index], field = lookup['field'])
    
    def add_1000_formula(self, start_row, put_col, ref_col, sheet_name, get_col='B'):
        for i in range(start_row, 1000):
            self.add_red_formula(i, put_col, i, ref_col, sheet_name, get_col)
    
    def add_red_formula(self, put_row, put_col, ref_row, ref_col, sheet_name, get_col='B'):
        try:
            get_col_i = int(get_col)
        except ValueError:
            pass
        else:
            get_col = openpyxl.cell.get_column_letter(get_col_i)
        c = self._ws.cell(row = put_row, column=put_col)
        ref_cell = self._ws.cell(row = ref_row, column=ref_col)
        c.value = '=IF(ISBLANK({address}),"",LOOKUP({address}, {sheet}!A$2:A$1000, {sheet}!{gc}$2:{gc}$1000))'.format(
                address = ref_cell.address, sheet = sheet_name, gc = get_col)
        self.set_red(c)
        return c
    
    def add_red_value(self, main_item, put_row, put_col, func=None, field=None):
        c = self._ws.cell(row = put_row, column=put_col)
        if func:
            c.value = getattr(main_item, func)()
        elif field:
            c.value = getattr(main_item, field)
        self.set_red(c)
                  
    def set_red(self, cell):
        color = getattr(openpyxl.style.Color, self.print_color, openpyxl.style.Color.RED)
        cell.style.font.color.index = color
        
class M2MExport(RedExtra):
    lookups = []
    
    def __init__(self, ws, firstcol):
        super(M2MExport, self).__init__(ws, firstcol)
        m2m_annotate = self.main_model.objects.all().annotate(other_count = db_models.Count(self.m2m_field_name))
        self._max_other = m2m_annotate.aggregate(db_models.Max('other_count'))['other_count__max']
        firstcol += len(self.lookups)
        self._m2m_red_columns = [i*2 + firstcol + 1 for i in range(self._max_other)]
        self._columns = [i*2 + firstcol for i in range(self._max_other)]
        
    def add_headings(self, row):
        super(M2MExport, self).add_headings(row)
        heads = [self.m2m_field_name for _ in range(self._max_other)]
        for (index, head) in enumerate(heads):
            c = self._ws.cell(row = row, column=self._columns[index])
            c.value = head
            c.style.font.bold = True
        for i in range(self._max_other):
            self.add_1000_formula(row + 1, self._m2m_red_columns[i], self._columns[i], self.m2m_lookup_sheet)
    
    def add_row(self, main_item, row):
        super(M2MExport, self).add_row(main_item, row)
        for (index, item) in enumerate(getattr(main_item, self.m2m_field_name).all()):
            c = self._ws.cell(row = row, column=self._columns[index])
            c.value = item.xl_id

class ImExBase:
    imex_fields = default_imex_fields
    field_headings = {}
    imex_order = 0
    imex_top_offset = 0
    import_edit_only = False
    import_sheet = True
    import_groups = []
    export_groups = []
    
    class ImportExtra(ImportExtra):
        pass
    
    class ExportExtra(ExportExtra):
        pass
    

def get_imex_app():
    import_str = settings.IMEX_APP + '.imex'
    return __import__(import_str)

def get_imex_groups():
    app = get_imex_app()
    import_groups = getattr(app.imex, 'IMPORT_GROUPS', (('all', 'Import All'),))
    export_groups = getattr(app.imex, 'EXPORT_GROUPS', (('all', 'Export All'),))
    return import_groups, export_groups
