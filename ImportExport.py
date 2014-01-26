from django.utils.encoding import smart_text
import openpyxl, re
from django.db import models
import inspect, traceback
from django.core.exceptions import ObjectDoesNotExist
import settings, os
from time import time

class _ImportExport:
    def get_models(self):
        imex_models = []
        import_str = settings.IMEX_APP + '.imex'
        try:
            app = __import__(import_str)
            perform_before_upload = None
            if hasattr(app.imex, 'PERFORM_BEFORE_UPLOAD'):
                perform_before_upload = getattr(app.imex, 'PERFORM_BEFORE_UPLOAD')
            for ob_name in dir(app.imex):
                ob = getattr(app.imex, ob_name)
                if inherits_from(ob, 'ImExBase'):
                    imex_models.append(ob)
            return sorted(imex_models, key=lambda mod: mod.imex_order), perform_before_upload
        except Exception, e:
            tb = traceback.format_exc().strip('\r\n')
            self._log('TRACEBACK:\n%s' % tb)
            msg = 'Error importing %s: %s' % (import_str, str(e))
            self._log(msg)
            raise Exception(msg)
        
    def raise_error(self, e):
        traceback.print_exc()
        tb = traceback.format_exc().strip('\r\n')
        self._log('TRACEBACK:')
        [self._log(line) for line in tb.split('\n')]
        if self._sheet:
            msg = 'Error on sheet %s, row %d: %r' % (self._sheet, self._row, e)
        else:
            msg = 'Error occurred outside sheet processing: %r' % e
        raise Exception(msg)

class ReadXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self.success = False
        self._sheet = None
        sheet_models, perform_before_upload = self.get_models()
        try:
            perform_before_upload(log)
        except Exception, e:
            self.raise_error(e)
        self._log('loading "%s"' % fname)
        self._wb = openpyxl.load_workbook(fname)
        self._sheet_names = self._wb.get_sheet_names()
        self._log('Worksheets: %r' % self._sheet_names)
        self._unichar_finder = re.compile(r'[^\x00-\xff]')
        self._row = 0
        try:
            for sheet_model in sheet_models:
                if sheet_model.import_sheet:
                    self._import_sheet(sheet_model, sheet_model.__name__)
        except Exception, e:
            self.raise_error(e)
        else:
            self.success = True
     
    def _import_sheet(self, sheet_model, sname, try_again = True):
        if sname not in self._sheet_names:
            self._log('"%s" is not a valid sheet name' % sname)
            if try_again:
                sname2 = '%s1' % sname
                self._log('trying "%s"' % sname2)
                self._import_sheet(sheet_model, sname2, False)
            return
        self._sheet_name = sname
        fields = sheet_model.imex_fields
        ws = self._wb.get_sheet_by_name(name = self._sheet_name)
        self._row = sheet_model.imex_top_offset
        headings = self._get_headings(ws)
        if None in headings:
            del headings[None]
#         self._log('column names: %r' % headings)
        extra = sheet_model.ImportExtra(ws, headings)
        self._row = self._row + 1
        import_count = 0
        for self._row in range(self._row, ws.get_highest_row()):
            xl_id = ws.cell(row=self._row, column=headings['xl_id']).value
            if not isinstance(xl_id, int):
                continue
            finds = sheet_model.main_model.objects.filter(xl_id = xl_id)
            if finds.count() == 1:
                main_item = finds[0]
            elif finds.count() == 0:
                if sheet_model.import_edit_only:
                    self._log('%s: model can only be edited not created via import, no item found on row %d'
                               % (self._sheet_name, self._row))
                    continue
                main_item = sheet_model.main_model()
            else:
                raise Exception('ERROR: already multiple items with the same xl_id(%d) in %s' % 
                                            (xl_id, self._sheet_name))
            for field in fields:
                value = ws.cell(row=self._row, column=headings[field]).value
#                 self._log('%s: %s' % (field, self._clean_string(value)))
                field_info = sheet_model.main_model._meta.get_field_by_name(field)[0]
                if isinstance(field_info, models.fields.related.ForeignKey):
                    if value is not None:
                        try:
                            value = field_info.rel.to.objects.get(xl_id = value)
                        except ObjectDoesNotExist:
                            raise Exception('ERROR: item with xl_id = %d does not exist in %s' % 
                                            (value, field_info.rel.to.__name__))
                else:
                    value = self._get_value(getattr(main_item, field), value)
                setattr(main_item, field, value)
            main_item.save()
            import_count += 1
            extra.get_row(main_item, self._row)
        self._log('imported %d %s' %  (import_count, sheet_model.main_model._meta.verbose_name_plural))
            
    def _get_headings(self, ws):
        headings={}
        for col_number in range(ws.get_highest_column()):
            name = ws.cell(row=self._row, column = col_number).value
            if name in headings:
                if isinstance(headings[name], list):
                    headings[name].append(col_number)
                else:
                    headings[name] = [headings[name], col_number]
            else:
                headings[name] = col_number
        return headings
    
    def _get_value(self, field, value):
        if value is None:
            return None
        elif isinstance(field, int) or isinstance(field, float):
            return value
        elif isinstance(field, str):
            return self._clean_string(value)
        else:
            return value
    
    def _clean_string(self, s):
            s = smart_text(s)
            return re.sub(self._unichar_finder, '', s)

class WriteXl(_ImportExport):
    def __init__(self, log):
        self._log = log
        self.success = False
        self._sheet = None
        self._wb = openpyxl.Workbook()
        self._row = 0
        self._delete_excess_sheets()
        export_models, _ = self.get_models()
        try:
            for export_model in export_models:
                self._log('Exporting data to %s' % export_model.__name__)
                start = time()
                export_count = self._write_model(export_model)
                self._log('    Exported %d items in %0.3fs' % (export_count, (time()-start)))
        except Exception, e:
            self.raise_error(e)
        else:
            self.fname = 'tmp.xlsx'
            #if settings.ON_SERVER:
                #fname = os.path.join(settings.SITE_ROOT,  self.fname)
            try:
                self._wb.save(filename = self.fname)
            except IOError:
                raise Exception('ERROR: writing to "%s" failed, file may be open' % self.fname)
            else:
                self._log('Writing output file')
                self.success = True
        
    def _write_model(self, sheet_model):
        ws = self._wb.create_sheet()
        self._sheet = sheet_model.__name__
        fields = sheet_model.imex_fields
        ws.title = self._sheet
        top_offset = sheet_model.imex_top_offset
        self._row = 0
        col = -1
        for (col, field) in enumerate(fields):
            c = ws.cell(row = top_offset, column=col)
            c.value = field
            c.style.font.bold = True
        for col_dim in ws.column_dimensions.values():
            col_dim.width = 15
            
        exportextra = sheet_model.ExportExtra(ws, col+1)
        exportextra.add_headings(top_offset)
            
        for (item_id, item) in enumerate(sheet_model.main_model.objects.all()):
            self._row = item_id + top_offset + 1
            for (col, field) in enumerate(fields):
                value = getattr(item, field)
                if hasattr(value, '__call__'):
                    value = value()
                if isinstance(value, models.Model):
                    value = value.xl_id
                elif isinstance(value, str) or isinstance(value, unicode):
                    value = value.strip('\r\n').replace('\r', '').replace('\n', ', ')
                if field =='xl_id' and value == -1:
                    value = item.id
                    item.xl_id =item.id
                    item.save()
                ws.cell(row = self._row, column=col).value = value
            exportextra.add_row(item, self._row)
        return item_id
                
    def _delete_excess_sheets(self):
        for sheet_name in self._wb.get_sheet_names():
            sheet = self._wb.get_sheet_by_name(sheet_name)
            self._wb.remove_sheet(sheet)
            
def inherits_from(child, parent_name):
    if inspect.isclass(child):
        if parent_name in [c.__name__ for c in inspect.getmro(child)[1:]]:
            return True
    return False
